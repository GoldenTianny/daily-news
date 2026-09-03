#!/usr/bin/env python3
# ETF_price_concensus_YYYYMMDD.xlsx -> 시장 데이터 DB(db/market/) 적재
#
# 신형(2026-08-29~, 당일 스냅샷): 시트 'ETF price'(ETF 수정주가), '수정주가, 목표주가'
#   (일반 종목 수정주가·목표주가·영업이익 컨센서스) — 기준일 1일치를 코드 단위로 병합
# 구형(시계열): 시트 '수정주가', 'consencus' — 원본에 있는 날짜 단위로 교체 병합
#
# 월별 파일로 저장하며, 내용이 기존과 같은 달은 건너뛰어 git 변경을 최소화.
# ('ETF raw' 시트는 tools/etf/build_data.py 담당이므로 여기서 다루지 않음)
# 사용법: python3 tools/market/build_market.py <xlsx> [YYYY-MM-DD(신형 기준일)]
import sys, os, re, numbers, datetime
import openpyxl
import duckdb
import pandas as pd

REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
DB = os.path.join(REPO, 'db', 'market')
SNAP_SHEET = '수정주가, 목표주가'   # 신형 스냅샷 시트 (존재 여부로 형식 판별)

SHEETS = [
    # (시트명, 하위 디렉터리, 값 컬럼명)
    ('수정주가', 'price', 'close'),
    ('consencus', 'consensus', 'target_price'),
]


def parse_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    s = str(v).strip()[:10]
    try:
        return datetime.date.fromisoformat(s)
    except ValueError:
        return None


def extract(ws, value_col):
    """와이드 시계열 시트 -> [(date, code, name, value)] (값 없는 셀은 제외)"""
    rows = list(ws.iter_rows(values_only=True))
    header = {str(r[0]).strip(): i for i, r in enumerate(rows[:14]) if r and r[0]}
    code_row = rows[header['Code']]
    name_row = rows[header['Name']]

    cols = []  # (col_idx, code, name)
    for c in range(1, len(code_row)):
        code = code_row[c]
        if code is None or str(code).strip() == '':
            continue
        name = name_row[c] if c < len(name_row) else None
        cols.append((c, str(code).strip(), str(name).strip() if name else ''))

    out = []
    for r in rows[header['D A T E'] + 1:]:
        d = parse_date(r[0]) if r and r[0] is not None else None
        if d is None:
            continue
        for c, code, name in cols:
            v = r[c] if c < len(r) else None
            if isinstance(v, numbers.Number):
                out.append((d, code, name, float(v)))
    return out


def write_monthly(records, sub, value_col, upsert=False):
    """월별 병합 저장. upsert=False: 날짜 단위 교체 / True: (날짜, 코드) 단위 갱신
    (스냅샷은 코드 단위 upsert — 그날 파일에 없는 종목의 기존 행을 지우지 않음)"""
    out_dir = os.path.join(DB, sub)
    os.makedirs(out_dir, exist_ok=True)
    df = pd.DataFrame(records, columns=['date', 'code', 'name', value_col])
    n_new = n_same = 0
    for ym, g in df.groupby(df['date'].map(lambda d: f'{d.year:04d}-{d.month:02d}')):
        dst = os.path.join(out_dir, ym + '.parquet')
        if os.path.exists(dst):
            old = pd.read_parquet(dst)
            old['date'] = pd.to_datetime(old['date']).dt.date
            if upsert:
                keys = set(zip(g['date'], g['code']))
                keep = old[[k not in keys for k in zip(old['date'], old['code'])]]
            else:
                keep = old[~old['date'].isin(set(g['date']))]
            g = pd.concat([keep, g])
            g = g.sort_values(['date', 'code']).reset_index(drop=True)
            if old.sort_values(['date', 'code']).reset_index(drop=True).equals(g):
                n_same += 1
                continue
        else:
            g = g.sort_values(['date', 'code']).reset_index(drop=True)
        con = duckdb.connect()
        con.execute(f"""
            COPY (SELECT CAST(date AS DATE) AS date, code, name,
                         CAST({value_col} AS DOUBLE) AS {value_col} FROM g)
            TO '{dst}' (FORMAT PARQUET, COMPRESSION SNAPPY)""")
        n_new += 1
    return len(df), df['code'].nunique(), n_new, n_same


def canonical_names():
    """가격 DB의 code -> name (스냅샷 원본의 '(주)OO' 표기를 기존 명칭으로 정규화)"""
    try:
        rows = duckdb.sql(
            f"SELECT code, any_value(name) AS name FROM '{os.path.join(DB, 'price', '*.parquet')}' GROUP BY code").fetchall()
        return {c: n for c, n in rows}
    except Exception:
        return {}


def parse_snapshot(ws, col, names, date_key):
    """스냅샷 시트의 열 하나 -> [(date, code, name, value)]"""
    out = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i < 13 or not r:
            continue
        code = str(r[0] or '').strip()
        v = r[col] if col < len(r) else None
        if not code.startswith('A') or not isinstance(v, numbers.Number):
            continue
        nm = names.get(code) or re.sub(r'^\(주\)|^㈜', '', str(r[1] or '').strip())
        out.append((date_key, code, nm, float(v)))
    return out


def build_snapshot(wb, date_key):
    """신형 스냅샷: ETF price + 수정주가/목표주가 -> price·consensus 1일 병합"""
    d = datetime.date.fromisoformat(date_key)
    names = canonical_names()
    price = []
    if 'ETF price' in wb.sheetnames:
        price += parse_snapshot(wb['ETF price'], 3, names, d)
    price += parse_snapshot(wb[SNAP_SHEET], 3, names, d)
    cons = parse_snapshot(wb[SNAP_SHEET], 4, names, d)
    for records, sub, vc in [(price, 'price', 'close'), (cons, 'consensus', 'target_price')]:
        n_rows, n_codes, n_new, n_same = write_monthly(records, sub, vc, upsert=True)
        print(f"OK  {date_key} 스냅샷 -> db/market/{sub}/: {n_rows:,}행 | "
              f"월 파일 {n_new}개 갱신, {n_same}개 동일(건너뜀)")


def build(xlsx_path, date_key=None):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if SNAP_SHEET in wb.sheetnames:                 # 신형 (당일 스냅샷)
        if not date_key:
            sys.exit("오류: 스냅샷 형식은 기준일(YYYY-MM-DD)이 필요합니다")
        build_snapshot(wb, date_key)
        return
    for sheet, sub, value_col in SHEETS:            # 구형 (시계열)
        if sheet not in wb.sheetnames:
            print(f"SKIP {sheet}: 시트 없음")
            continue
        records = extract(wb[sheet], value_col)
        n_rows, n_codes, n_new, n_same = write_monthly(records, sub, value_col)
        print(f"OK  {sheet} -> db/market/{sub}/: {n_rows:,}행, 종목 {n_codes:,}개 | "
              f"월 파일 {n_new}개 갱신, {n_same}개 동일(건너뜀)")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        sys.exit("사용법: python3 tools/market/build_market.py <ETF_price_concensus_YYYYMMDD.xlsx> [YYYY-MM-DD]")
    build(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
