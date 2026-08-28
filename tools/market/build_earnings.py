#!/usr/bin/env python3
"""concensus_for_db*.xlsx -> 영업이익 실적·컨센서스 DB (db/market/earnings/)

원본 시트 구조:
  - 'fixed'            : 확정 영업이익 (M121500) — Period 행의 'YYYYAS'(연간) 컬럼만 사용.
                         4분기 실적이 확정되면 해당 연도 AS 컬럼이 채워져 시퀀스가 자동 연장됨
  - 'YYYY annual margin': 해당 회계연도 연간 영업이익 컨센서스(E121500)의 일별 추이
                         (Base Date 행이 기준일. 'CPD-1TD' 같은 비날짜 컬럼은 제외)

산출물:
  - db/market/earnings/annual.parquet          (code, name, fy, op, src 'A'확정/'E'컨센서스 최신)
  - db/market/earnings/consensus/YYYY-MM.parquet (date, code, name, fy, op) — 월별, 날짜 단위 병합
단위: 천원(Local thousand). 종목명은 가격 DB(db/market/price)의 명칭으로 정규화.
사용법: python3 tools/market/build_earnings.py <concensus_for_db*.xlsx>
"""
import sys, os, re, numbers
import openpyxl
import duckdb
import pandas as pd

REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
OUT = os.path.join(REPO, 'db', 'market', 'earnings')
PRICE = os.path.join(REPO, 'db', 'market', 'price', '*.parquet')


def canonical_names():
    """가격 DB의 code -> name (검색기 종목명과 동일 체계)"""
    try:
        rows = duckdb.sql(
            f"SELECT code, any_value(name) AS name FROM '{PRICE}' GROUP BY code").fetchall()
        return {c: n for c, n in rows}
    except Exception:
        return {}


def parse_fixed(ws, names):
    """확정 영업이익 -> (연간 [(code, name, fy, op)], 분기 [(code, name, fy, q, op)])"""
    rows = list(ws.iter_rows(values_only=True))
    periods = rows[9]
    fy_cols = [(i, int(str(p)[:4])) for i, p in enumerate(periods)
               if re.fullmatch(r'\d{4}AS', str(p or ''))]
    q_cols = [(i, int(str(p)[:4]), int(str(p)[4:6]) // 3) for i, p in enumerate(periods)
              if re.fullmatch(r'\d{6}', str(p or ''))]        # YYYYMM(3/6/9/12) -> 분기
    annual, quarterly = [], []
    for r in rows[13:]:
        code = str(r[0] or '').strip()
        if not code.startswith('A'):
            continue
        nm = names.get(code) or str(r[1] or '').strip()
        for i, fy in fy_cols:
            v = r[i] if i < len(r) else None
            if isinstance(v, numbers.Number):
                annual.append((code, nm, fy, float(v)))
        for i, fy, q in q_cols:
            v = r[i] if i < len(r) else None
            if isinstance(v, numbers.Number):
                quarterly.append((code, nm, fy, q, float(v)))
    return annual, quarterly


def parse_margin(ws, names):
    """연간 컨센서스 일별 추이 -> (fy, [(date, code, name, fy, op)])"""
    rows = list(ws.iter_rows(values_only=True))
    m = re.search(r'(\d{4})AS', str(rows[9][3] or ''))
    if not m:
        return None, []
    fy = int(m.group(1))
    base = rows[11]
    date_cols = []
    for i in range(3, len(base)):
        s = str(base[i] or '').strip()
        if re.fullmatch(r'\d{8}', s):                 # 'CPD-1TD' 등 비날짜 컬럼 제외
            date_cols.append((i, f'{s[:4]}-{s[4:6]}-{s[6:]}'))
    out = []
    for r in rows[13:]:
        code = str(r[0] or '').strip()
        if not code.startswith('A'):
            continue
        nm = names.get(code) or str(r[1] or '').strip()
        for i, d in date_cols:
            v = r[i] if i < len(r) else None
            if isinstance(v, numbers.Number):
                out.append((d, code, nm, fy, float(v)))
    return fy, out


def norm(d):
    """비교용 정규화 (parquet 왕복 시 dtype이 달라져도 동일 판정되도록)"""
    d = d.copy()
    d['date'] = d['date'].astype(str).str[:10]
    d['fy'] = d['fy'].astype('int64')
    d['op'] = d['op'].astype('float64')
    return d.sort_values(['date', 'fy', 'code']).reset_index(drop=True)


def write_monthly(df):
    """consensus/YYYY-MM.parquet — 원본에 있는 날짜만 교체, 나머지 유지"""
    out_dir = os.path.join(OUT, 'consensus')
    os.makedirs(out_dir, exist_ok=True)
    n_new = n_same = 0
    con = duckdb.connect()
    for ym, g in df.groupby(df['date'].str[:7]):
        dst = os.path.join(out_dir, ym + '.parquet')
        if os.path.exists(dst):
            old = pd.read_parquet(dst)
            old['date'] = old['date'].astype(str).str[:10]
            g = pd.concat([old[~old['date'].isin(set(g['date']))], g])
        g = norm(g)
        if os.path.exists(dst) and norm(pd.read_parquet(dst)).equals(g):
            n_same += 1
            continue
        con.execute(f"""
            COPY (SELECT CAST(date AS DATE) AS date, code, name,
                         CAST(fy AS SMALLINT) AS fy, CAST(op AS DOUBLE) AS op FROM g)
            TO '{dst}' (FORMAT PARQUET, COMPRESSION SNAPPY)""")
        n_new += 1
    return n_new, n_same


def build(xlsx_path):
    names = canonical_names()
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    actual, quarterly = [], []
    if 'fixed' in wb.sheetnames:
        actual, quarterly = parse_fixed(wb['fixed'], names)
        print(f"OK  확정 실적: 연간 {len(actual):,}행 "
              f"({min(a[2] for a in actual)}~{max(a[2] for a in actual)}) · 분기 {len(quarterly):,}행")

    cons_rows = []
    latest = {}   # (code, fy) -> (name, date, op)
    for sn in wb.sheetnames:
        if 'annual margin' not in sn:
            continue
        fy, rows = parse_margin(wb[sn], names)
        if fy is None:
            continue
        cons_rows += rows
        for d, code, nm, f, v in rows:
            k = (code, f)
            if k not in latest or d > latest[k][1]:
                latest[k] = (nm, d, v)
        print(f"OK  {fy} 컨센서스 추이: {len(rows):,}행")

    os.makedirs(OUT, exist_ok=True)

    # annual.parquet: 확정(A) + 확정 없는 연도의 최신 컨센서스(E)
    confirmed = {(c, fy) for c, _, fy, _ in actual}
    ann = [(c, n, fy, v, 'A') for c, n, fy, v in actual]
    ann += [(c, nm, fy, v, 'E') for (c, fy), (nm, d, v) in latest.items()
            if (c, fy) not in confirmed]
    df = pd.DataFrame(ann, columns=['code', 'name', 'fy', 'op', 'src'])
    df = df.sort_values(['code', 'fy']).reset_index(drop=True)
    dst = os.path.join(OUT, 'annual.parquet')
    duckdb.connect().execute(f"""
        COPY (SELECT code, name, CAST(fy AS SMALLINT) AS fy,
                     CAST(op AS DOUBLE) AS op, src FROM df)
        TO '{dst}' (FORMAT PARQUET, COMPRESSION SNAPPY)""")
    print(f"OK  annual.parquet: {len(df):,}행, 기업 {df['code'].nunique():,}개 "
          f"({os.path.getsize(dst) // 1024} KB)")

    if quarterly:
        qdf = pd.DataFrame(quarterly, columns=['code', 'name', 'fy', 'q', 'op'])
        qdf = qdf.sort_values(['code', 'fy', 'q']).reset_index(drop=True)
        qdst = os.path.join(OUT, 'quarterly.parquet')
        duckdb.connect().execute(f"""
            COPY (SELECT code, name, CAST(fy AS SMALLINT) AS fy, CAST(q AS TINYINT) AS q,
                         CAST(op AS DOUBLE) AS op FROM qdf)
            TO '{qdst}' (FORMAT PARQUET, COMPRESSION SNAPPY)""")
        print(f"OK  quarterly.parquet: {len(qdf):,}행 ({os.path.getsize(qdst) // 1024} KB)")

    if cons_rows:
        cdf = pd.DataFrame(cons_rows, columns=['date', 'code', 'name', 'fy', 'op'])
        n_new, n_same = write_monthly(cdf)
        print(f"OK  consensus/: {len(cdf):,}행 | 월 파일 {n_new}개 갱신, {n_same}개 동일")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        sys.exit('사용법: python3 tools/market/build_earnings.py <concensus_for_db*.xlsx>')
    build(sys.argv[1])
