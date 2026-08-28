#!/usr/bin/env python3
"""일일 원본 엑셀 하나로 ETF·수정주가·컨센서스·RS 등급 DB를 한 번에 갱신.

  python3 tools/ingest_daily.py ~/Downloads/ETF_price_concensus_20260818.xlsx

- 기준일은 'ETF raw' 시트의 Date 셀(CPD [YYYYMMDD] / CPD-1TD [YYYYMMDD])에서 자동 인식
  (인식 실패 시 두 번째 인자로 YYYY-MM-DD 직접 지정)
- 실행 순서: build_data(ETF 보유내역) -> build_market(수정주가·컨센서스) -> build_rs(RS 등급)
  -> build_high52(52주 신고가 돌파 분석, 스터디)
- 어느 디렉터리에서 실행해도 저장소 기준 경로로 동작
"""
import sys, os, re, datetime, numbers
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
sys.path.insert(0, os.path.join(HERE, 'etf'))
sys.path.insert(0, os.path.join(HERE, 'market'))
sys.path.insert(0, os.path.join(HERE, 'study'))
import build_data, build_market, build_rs, build_high52


def detect_date(xlsx_path):
    """기준일(종가 기준일, YYYY-MM-DD) 인식.

    1순위: '수정주가' 시트에서 실제 가격이 채워진 마지막 날짜.
      CPD 태그는 장 시작 전에 받으면 다음 거래일을 가리켜 하루 밀리므로
      (예: 8/28 07시 다운로드 → CPD [20260828], 실제 데이터는 8/27 종가),
      값이 존재하는 마지막 날짜를 종가 기준일로 본다.
    2순위: 'ETF raw' 시트의 Date 셀 CPD 태그 (수정주가 시트가 없을 때).
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    if '수정주가' in wb.sheetnames:
        last = None
        for row in wb['수정주가'].iter_rows(values_only=True):
            d = row[0] if row else None
            if isinstance(d, datetime.datetime):
                if any(isinstance(v, numbers.Number) for v in row[1:]):
                    last = d.date()
        if last:
            return last.isoformat()

    ws = wb['ETF raw'] if 'ETF raw' in wb.sheetnames else wb[wb.sheetnames[0]]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 8:
            break
        for cell in row[:4]:
            m = re.search(r'\[(\d{8})\]', str(cell or ''))
            if m:
                d = m.group(1)
                return f'{d[:4]}-{d[4:6]}-{d[6:]}'
    return None


def main():
    if len(sys.argv) < 2:
        sys.exit('사용법: python3 tools/ingest_daily.py <ETF_price_concensus_YYYYMMDD.xlsx> [YYYY-MM-DD]')
    xlsx = sys.argv[1]

    # 실적·컨센서스 파일(concensus_for_db*.xlsx)이면 earnings 적재만 수행
    sheets = openpyxl.load_workbook(xlsx, read_only=True).sheetnames
    if any('annual margin' in s for s in sheets):
        import build_earnings
        print(f'== 영업이익 실적·컨센서스 파일 · {os.path.basename(xlsx)}')
        build_earnings.build(xlsx)
        print('== 완료')
        return

    date_key = sys.argv[2] if len(sys.argv) > 2 else detect_date(xlsx)
    if not date_key:
        sys.exit('오류: 기준일을 인식하지 못했습니다. 두 번째 인자로 YYYY-MM-DD를 지정해주세요.')

    print(f'== 기준일 {date_key} · {os.path.basename(xlsx)}')
    print('[1/4] ETF 보유내역')
    build_data.build(xlsx, date_key, os.path.join(REPO, 'tools', 'etf', 'data'))
    print('[2/4] 수정주가 · 컨센서스')
    build_market.build(xlsx)
    print('[3/4] RS 등급')
    build_rs.build()
    print('[4/4] 52주 신고가 돌파 분석')
    build_high52.build()
    print('== 완료')


if __name__ == '__main__':
    main()
